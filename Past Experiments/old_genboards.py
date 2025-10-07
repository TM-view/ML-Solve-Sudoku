# genboards.py
import random
import numpy as np
from openpyxl import Workbook

# สร้าง array แทน list ทั้งหมด
ANS = np.zeros((9, 9), dtype=int)
max_loss = np.array([20, 40, 60], dtype=int)
max_loss_set = np.array([3, 5, 8], dtype=int)
rate_diff = np.array([6, 4, 2], dtype=int)

def fill_board():
    for row in range(9):
        for col in range(9):
            if ANS[row, col] == 0:
                nums = np.arange(1, 10)  # แทน list(range(1,10))
                np.random.shuffle(nums)   # สุ่มเลข
                for num in nums:
                    if is_valid(ANS, row, col, num):
                        ANS[row, col] = num
                        if fill_board():
                            return True
                        ANS[row, col] = 0
                return False
    return True

def is_valid(board, row, col, num):
    # ตรวจ row และ col
    if num in board[row, :] or num in board[:, col]:
        return False
    # ตรวจ 3x3 box
    start_row, start_col = 3 * (row // 3), 3 * (col // 3)
    if num in board[start_row:start_row+3, start_col:start_col+3]:
        return False
    return True

def print_board(SHOW):
    all_loss = 0
    each_loss = 0
    difficult = 0
    for i in range(9):
        for j in range(9):
            if random.randint(1, 11) > rate_diff[difficult]:
                SHOW[i, j] = 0
                all_loss += 1
                each_loss += 1
                if all_loss >= max_loss[difficult]:
                    return
                if each_loss >= max_loss_set[difficult]:
                    break
        each_loss = 0

def main(num=1):
    # สร้าง NumPy array สำหรับเก็บทั้งหมด (num boards × 2 × 9 × 9)
    OVERALL = np.zeros((num, 2, 9, 9), dtype=int)

    for i in range(num):
        ANS[:] = 0
        fill_board()
        SHOW = ANS.copy()
        print_board(SHOW)
        
        # เก็บข้อมูลลง array 3D
        OVERALL[i, 0] = SHOW     # PLAY
        OVERALL[i, 1] = ANS      # ANSWER

    # เขียน Excel
    wb = Workbook()
    ws = wb.active

    row = 1
    for i in range(num):
        ws[f"A{row}"] = str(OVERALL[i, 0].tolist())  # PLAY
        ws[f"B{row}"] = str(OVERALL[i, 1].tolist())  # ANSWER
        row += 1

    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 100

    wb.save("sudoku_dataset.xlsx")
    print("บันทึก Excel เรียบร้อยแล้ว!")

# call_main
main(2)
