from openpyxl import Workbook

# ข้อมูลหลายบอร์ด Sudoku
boards = [
    [
        [0, 0, 0, 0, 2, 4, 5, 0, 0],
        [0, 0, 0, 0, 0, 0, 0, 0, 0],
        [5, 2, 0, 0, 0, 7, 0, 0, 4],
        [0, 0, 0, 2, 0, 0, 4, 9, 0],
        [3, 0, 0, 0, 0, 0, 0, 0, 5],
        [0, 0, 0, 5, 7, 6, 2, 0, 3],
        [2, 0, 0, 0, 0, 0, 8, 5, 0],
        [0, 0, 0, 0, 0, 0, 0, 0, 0],
        [0, 8, 3, 0, 0, 0, 7, 0, 2]
    ],
    [
        [1, 0, 0, 0, 0, 0, 0, 0, 9],
        [0, 0, 0, 0, 3, 0, 0, 0, 0],
        [0, 2, 0, 5, 0, 0, 0, 0, 0],
        [0, 0, 3, 0, 0, 2, 0, 4, 0],
        [0, 0, 0, 7, 0, 8, 0, 0, 0],
        [0, 6, 0, 4, 0, 0, 3, 0, 0],
        [0, 0, 0, 0, 0, 3, 0, 1, 0],
        [0, 0, 0, 0, 5, 0, 0, 0, 0],
        [8, 0, 0, 0, 0, 0, 0, 0, 7]
    ]
]

# สร้าง workbook
wb = Workbook()
ws = wb.active

# ใส่ข้อมูลหลายบอร์ด
for i, board in enumerate(boards, start=1):
    ws[f"A{i}"] = str(board)        # ใส่ board ลง column A
    ws[f"B{i}"] = f"Board {i}"      # ใส่ label ลง column B

# ปรับ column width ให้เห็นเต็ม
ws.column_dimensions["A"].width = 150
ws.column_dimensions["B"].width = 100

# บันทึกไฟล์
wb.save("multiple_sudoku.xlsx")

print("สร้างไฟล์ Excel หลาย row เรียบร้อยแล้ว!")
