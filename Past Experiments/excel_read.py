from openpyxl import load_workbook
import ast  # แปลง string เป็น list safely

# โหลดไฟล์ Excel
wb = load_workbook("multiple_sudoku.xlsx")
ws = wb.active

boards = []      # เก็บ board ทั้งหมด
labels = []      # เก็บ label ใน column B

# ลูปอ่านทุก row ที่มีข้อมูล
for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
    cell_a, cell_b = row
    if cell_a is not None:
        board = ast.literal_eval(cell_a)  # แปลง string -> list
        boards.append(board)
        labels.append(cell_b)

# แสดงผล
for i, (board, label) in enumerate(zip(boards, labels), start=1):
    print(f"\nBoard {i} ({label}):")
    for r in board:
        print(r)

# ตัวอย่างเข้าถึง element
print("\nตัวอย่างเข้าถึง Board 1 แถวแรก คอลัมน์แรก:", boards[0][0][0])
